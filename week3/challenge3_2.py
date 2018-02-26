import os
from datetime import datetime
from openpyxl import load_workbook, Workbook

wb = load_workbook('courses.xlsx')
student_sheet = wb['students']
students = []

def combine():
	time_sheet = wb['time']
	time_dict = {}

	for i in range(1, time_sheet.max_row + 1):
		course_name = time_sheet[i][1].value
		time_dict[course_name] = time_sheet[i][2].value

	for i in range(1, student_sheet.max_row + 1):
		stu = [x.value for x in student_sheet[i]]
		students.append(stu)

	for stu in students:
		time = time_dict.get(stu[1], None)
		stu.append(time)

	combine_sheet = wb.create_sheet('combine')

	for row in students:
		for col in row:
			row_index = students.index(row)
			col_index = row.index(col)
			#combine_sheet[row_index+1][col_index+1] = students[row_index][col_index]
			combine_sheet.cell(row=row_index+1, column=col_index+1).value = students[row_index][col_index]
	os.rename('courses.xlsx', 'courses' + datetime.strftime(datetime.now(), '%Y%m%d%H%M%S') + '.xlsx')
	wb.save('courses.xlsx')

def split():
	year_dict = {}
	for row in students[1:]:
		year_of_row = row[0].year

		if year_dict.get(year_of_row) is None:
			year_dict[year_of_row] = []

		year_dict[year_of_row].append(row)

	for year in year_dict.keys():
		year_wb = Workbook()
		year_sheet = year_wb.active
		year_sheet.title = str(year)
		year_sheet.append(students[0])
		for item in year_dict[year]:
			year_sheet.append(item)

		year_wb.save(str(year) + '.xlsx')
		year_wb.close()



if __name__ == '__main__':
	combine()
	split()